from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


MONTH_SHEETS = ("AGOSTO", "SETEMBRO", "OUTUBRO", "JANEIRO", "FEVEREIRO")
TIME_PREFIX_RE = re.compile(r"^(?P<time>\d{1,2}h(?:\d{2})?)\s*-\s*(?P<body>.+)$", re.IGNORECASE)
SINGLE_BOX_RE = re.compile(r"^(T|G|D|Q|R)(?:\s*\([^)]*\)|\s+C/.*)?$", re.IGNORECASE)
MIXED_BOX_RE = re.compile(r"^(MG|MD|MQ|MR)(?:\s*\([^)]*\))?$", re.IGNORECASE)
EXPLICIT_BOX_RE = re.compile(r"^\s*\d+\s*[TGDQR](?:\s*\+\s*\d+\s*[TGDQR])+\s*$", re.IGNORECASE)
EXPLICIT_PART_RE = re.compile(r"(\d+)\s*([TGDQR])", re.IGNORECASE)
IMPORT_TAG = "[IMPORTADO_PLANILHA_LEGADA]"
LOCAL_TZ = ZoneInfo("America/Sao_Paulo")

ORDER_BOX_UNITS = 7
BOX_PRICE_CUSTOM = 52.0
BOX_PRICE_TRADITIONAL = 40.0
BOX_PRICE_MIXED_GOIABADA = 45.0
BOX_PRICE_MIXED_OTHER = 47.0
BOX_PRICE_GOIABADA = 50.0
HISTORICAL_CUSTOM_BOX_CODE = "__HIST_CUSTOM_BOX__"
HISTORICAL_CUSTOM_BOX_NAME = "Caixa historica sem composicao"
HISTORICAL_PRODUCT_CATEGORY = "Historico"
HISTORICAL_CUSTOM_BOX_HINTS = {
    "SABORES",
    "M",
    "QD",
    "QG",
    "Q + G",
    "R + D",
    "2 CADA",
    "2 DE CADA",
    "MG + 1D",
    "MG - COLOCAR 1Q",
    "Q + 4 EXTRAS G",
    "QUEIJO + GOIABADA",
}
IGNORED_INVALID_REASONS = {
    "CAIXA P",
}

MIXED_ALIAS_COUNTS: dict[str, dict[str, int]] = {
    "TRADICIONAL + GOIABADA": {"T": 4, "G": 3},
    "TRADICIONAL + DOCE DE LEITE": {"T": 4, "D": 3},
    "TRADICIONAL + QUEIJO": {"T": 4, "Q": 3},
    "TRADICIONAL + REQUEIJAO": {"T": 4, "R": 3},
}


@dataclass
class LegacyRow:
    sheet: str
    iso_date: str
    row: int
    column: int
    raw_value: str
    customer_name: str
    customer_key: str
    time_hint: str | None
    order_hint: str | None


@dataclass
class ParsedBox:
    raw_hint: str
    normalized_hint: str
    counts: dict[str, int]
    box_price: float
    parse_kind: str
    note: str | None = None


@dataclass
class ImportGroup:
    sheet: str
    iso_date: str
    customer_name: str
    customer_key: str
    time_hint: str | None
    rows: list[LegacyRow]


@dataclass
class ResolutionEntry:
    import_key: str
    resolution_rows: list[str]
    resolution_notes: str | None = None


def normalize_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_lookup(value: str | None) -> str:
    normalized = unicodedata.normalize("NFD", normalize_text(value))
    normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return normalized.upper()


def split_legacy_entry(raw: str) -> tuple[str, str | None, str | None]:
    normalized = normalize_text(raw)
    if not normalized:
        return "", None, None

    time_hint = None
    body = normalized
    time_match = TIME_PREFIX_RE.match(normalized)
    if time_match:
        time_hint = normalize_text(time_match.group("time")).lower()
        body = normalize_text(time_match.group("body"))

    if " - " in body:
        customer_name, order_hint = body.split(" - ", 1)
        return normalize_text(customer_name), time_hint, normalize_text(order_hint)

    return body, time_hint, None


def resolve_box_price(counts: dict[str, int]) -> float:
    total_units = sum(counts.values())
    if total_units != ORDER_BOX_UNITS:
        raise ValueError(f"caixa invalida com {total_units} unidade(s)")

    if counts == {"T": ORDER_BOX_UNITS}:
        return BOX_PRICE_TRADITIONAL
    if counts == {"G": ORDER_BOX_UNITS}:
        return BOX_PRICE_GOIABADA
    if counts in ({"D": ORDER_BOX_UNITS}, {"Q": ORDER_BOX_UNITS}, {"R": ORDER_BOX_UNITS}):
        return BOX_PRICE_CUSTOM
    if counts == {"T": 4, "G": 3}:
        return BOX_PRICE_MIXED_GOIABADA
    if counts in ({"T": 4, "D": 3}, {"T": 4, "Q": 3}, {"T": 4, "R": 3}):
        return BOX_PRICE_MIXED_OTHER
    return BOX_PRICE_CUSTOM


def parse_box_hint(order_hint: str | None) -> ParsedBox | None:
    if not order_hint:
        return None

    raw_hint = normalize_text(order_hint)
    normalized_hint = normalize_lookup(raw_hint)
    if not normalized_hint:
        return None

    if normalized_hint.startswith("(") and normalized_hint.endswith(")"):
        inner = normalized_hint[1:-1].strip()
        if inner:
            normalized_hint = inner

    if normalized_hint in HISTORICAL_CUSTOM_BOX_HINTS:
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts={HISTORICAL_CUSTOM_BOX_CODE: 1},
            box_price=BOX_PRICE_CUSTOM,
            parse_kind="historical-opaque",
            note="caixa historica sem composicao detalhada",
        )

    alias_counts = MIXED_ALIAS_COUNTS.get(normalized_hint)
    if alias_counts is not None:
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts=alias_counts,
            box_price=resolve_box_price(alias_counts),
            parse_kind="mixed-alias",
            note=raw_hint,
        )

    single_match = SINGLE_BOX_RE.match(normalized_hint)
    if single_match:
        code = single_match.group(1).upper()
        counts = {code: ORDER_BOX_UNITS}
        note = raw_hint if normalized_hint not in {"T", "G", "D", "Q", "R"} else None
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts=counts,
            box_price=resolve_box_price(counts),
            parse_kind="single",
            note=note,
        )

    if normalized_hint in {"T COM EV", "T SEM LAC", "SEM LAC G"}:
        code = "G" if normalized_hint == "SEM LAC G" else "T"
        counts = {code: ORDER_BOX_UNITS}
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts=counts,
            box_price=resolve_box_price(counts),
            parse_kind="single-note",
            note=raw_hint,
        )

    mixed_match = MIXED_BOX_RE.match(normalized_hint)
    if mixed_match:
        code = mixed_match.group(1).upper()
        counts = {"T": 4, code[1]: 3}
        note = raw_hint if raw_hint.upper() != code else None
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts=counts,
            box_price=resolve_box_price(counts),
            parse_kind="mixed",
            note=note,
        )

    if EXPLICIT_BOX_RE.match(normalized_hint):
        counts: Counter[str] = Counter()
        for quantity, code in EXPLICIT_PART_RE.findall(normalized_hint):
            counts[code.upper()] += int(quantity)
        counts_dict = dict(counts)
        return ParsedBox(
            raw_hint=raw_hint,
            normalized_hint=normalized_hint,
            counts=counts_dict,
            box_price=resolve_box_price(counts_dict),
            parse_kind="explicit",
        )

    return None


def parse_order_datetime(iso_date: str, time_hint: str | None) -> str:
    base_date = datetime.fromisoformat(iso_date)
    if time_hint:
        hours_text, _, minutes_text = time_hint.partition("h")
        hour = int(hours_text)
        minute = int(minutes_text or "0")
    else:
        hour = 12
        minute = 0

    local_dt = datetime(
        base_date.year,
        base_date.month,
        base_date.day,
        hour,
        minute,
        0,
        0,
        tzinfo=LOCAL_TZ,
    )
    return local_dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def load_rows(workbook_path: Path) -> list[LegacyRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    rows: list[LegacyRow] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name not in MONTH_SHEETS:
            continue

        ws = workbook[sheet_name]
        for column in range(1, ws.max_column + 1):
            date_value = ws.cell(1, column).value
            if date_value is None:
                continue

            if hasattr(date_value, "date"):
                iso_date = date_value.date().isoformat()
            else:
                iso_date = normalize_text(str(date_value))

            for row_index in range(2, ws.max_row + 1):
                cell_value = ws.cell(row_index, column).value
                if cell_value is None:
                    continue

                raw_value = normalize_text(str(cell_value))
                if not raw_value:
                    continue

                customer_name, time_hint, order_hint = split_legacy_entry(raw_value)
                rows.append(
                    LegacyRow(
                        sheet=sheet_name,
                        iso_date=iso_date,
                        row=row_index,
                        column=column,
                        raw_value=raw_value,
                        customer_name=customer_name,
                        customer_key=normalize_lookup(customer_name),
                        time_hint=time_hint,
                        order_hint=order_hint,
                    )
                )

    return rows


def group_rows(rows: list[LegacyRow]) -> list[ImportGroup]:
    grouped: dict[tuple[str, str, str], list[LegacyRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.iso_date, row.time_hint or "", row.customer_key)].append(row)

    groups: list[ImportGroup] = []
    for _, group_rows_list in grouped.items():
        ordered = sorted(group_rows_list, key=lambda row: (row.column, row.row))
        first = ordered[0]
        groups.append(
            ImportGroup(
                sheet=first.sheet,
                iso_date=first.iso_date,
                customer_name=first.customer_name,
                customer_key=first.customer_key,
                time_hint=first.time_hint,
                rows=ordered,
            )
        )

    groups.sort(key=lambda group: (group.iso_date, group.time_hint or "", group.customer_name))
    return groups


def build_import_plan(groups: list[ImportGroup]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], Counter[str]]:
    return build_import_plan_with_resolutions(groups, {})


def build_import_plan_with_resolutions(
    groups: list[ImportGroup],
    resolutions: dict[str, ResolutionEntry],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], Counter[str], Counter[str]]:
    importable: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    ignored: list[dict[str, Any]] = []
    reasons = Counter()
    ignored_reasons = Counter()

    for group in groups:
        import_key = f"{group.iso_date}|{group.time_hint or 'sem-hora'}|{group.customer_key}"
        parsed_boxes: list[ParsedBox] = []
        invalid_rows: list[dict[str, Any]] = []

        for row in group.rows:
            parsed = parse_box_hint(row.order_hint)
            if parsed is None:
                reason = normalize_lookup(row.order_hint or "sem composicao") or "SEM_COMPOSICAO"
                invalid_rows.append(
                    {
                        "raw_value": row.raw_value,
                        "reason": reason,
                        "row": row.row,
                        "column": row.column,
                    }
                )
                continue
            parsed_boxes.append(parsed)

        ignored_rows = [row for row in invalid_rows if row.get("reason") in IGNORED_INVALID_REASONS]
        invalid_rows = [row for row in invalid_rows if row.get("reason") not in IGNORED_INVALID_REASONS]

        for row in ignored_rows:
            ignored_reasons[row["reason"]] += 1

        resolution = resolutions.get(import_key)
        if resolution and resolution.resolution_rows:
            resolution_invalid = False
            parsed_boxes = []
            invalid_rows = []
            for index, raw_resolution in enumerate(resolution.resolution_rows, start=1):
                parsed = parse_box_hint(raw_resolution)
                if parsed is None:
                    reason = f"RESOLUCAO_INVALIDA:{normalize_lookup(raw_resolution) or 'VAZIA'}"
                    invalid_rows.append(
                        {
                            "raw_value": raw_resolution,
                            "reason": reason,
                            "row": index,
                            "column": 0,
                        }
                    )
                    resolution_invalid = True
                    continue
                parsed_boxes.append(parsed)
            if resolution_invalid:
                skipped.append(
                    {
                        "sheet": group.sheet,
                        "iso_date": group.iso_date,
                        "customer_name": group.customer_name,
                        "time_hint": group.time_hint,
                        "import_key": import_key,
                        "raw_rows": [row.raw_value for row in group.rows],
                        "invalid_rows": invalid_rows,
                        "resolution_rows": resolution.resolution_rows,
                    }
                )
                continue

        if invalid_rows:
            for row in invalid_rows:
                reasons[row["reason"]] += 1
            skipped.append(
                {
                    "sheet": group.sheet,
                    "iso_date": group.iso_date,
                    "customer_name": group.customer_name,
                    "time_hint": group.time_hint,
                    "import_key": import_key,
                    "raw_rows": [row.raw_value for row in group.rows],
                    "invalid_rows": invalid_rows,
                }
            )
            continue

        if not parsed_boxes and ignored_rows:
            ignored.append(
                {
                    "sheet": group.sheet,
                    "iso_date": group.iso_date,
                    "customer_name": group.customer_name,
                    "time_hint": group.time_hint,
                    "import_key": import_key,
                    "raw_rows": [row.raw_value for row in group.rows],
                    "ignored_rows": ignored_rows,
                }
            )
            continue

        aggregated_counts: Counter[str] = Counter()
        notes_parts: list[str] = []
        subtotal = 0.0
        for box in parsed_boxes:
            aggregated_counts.update(box.counts)
            subtotal += box.box_price
            notes_parts.append(box.raw_hint)

        scheduled_at = parse_order_datetime(group.iso_date, group.time_hint)
        resolution_notes = resolution.resolution_notes if resolution else None
        notes_tail = f" resolucao={resolution_notes}" if resolution_notes else ""
        ignored_tail = ""
        if ignored_rows:
            ignored_tail = f" ignorados={' | '.join(row['raw_value'] for row in ignored_rows)}"
        importable.append(
            {
                "sheet": group.sheet,
                "iso_date": group.iso_date,
                "customer_name": group.customer_name,
                "customer_key": group.customer_key,
                "time_hint": group.time_hint,
                "scheduled_at": scheduled_at,
                "counts": dict(aggregated_counts),
                "subtotal": round(subtotal, 2),
                "box_count": len(parsed_boxes),
                "raw_rows": [row.raw_value for row in group.rows],
                "import_key": import_key,
                "notes": f"{IMPORT_TAG} key={import_key} origem={group.sheet} caixas={' | '.join(notes_parts)}{ignored_tail}{notes_tail}",
                "resolved": bool(resolution and resolution.resolution_rows),
            }
        )

    return importable, skipped, ignored, reasons, ignored_reasons


def load_resolutions(path: Path) -> dict[str, ResolutionEntry]:
    if not path.exists():
        return {}

    resolutions: dict[str, ResolutionEntry] = {}
    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            import_key = normalize_text(row.get("import_key"))
            if not import_key:
                continue
            resolution_rows = [
                normalize_text(part)
                for part in re.split(r"\s*\|\s*", row.get("resolution_rows") or "")
                if normalize_text(part)
            ]
            resolution_notes = normalize_text(row.get("resolution_notes")) or None
            resolutions[import_key] = ResolutionEntry(
                import_key=import_key,
                resolution_rows=resolution_rows,
                resolution_notes=resolution_notes,
            )
    return resolutions


def fetch_products(connection: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    rows = connection.execute("select id, name, price, active from Product order by id asc").fetchall()
    product_by_code: dict[str, dict[str, Any]] = {}

    for row in rows:
        normalized_name = normalize_lookup(row["name"])
        code = None
        if "TRADICIONAL" in normalized_name:
            code = "T"
        elif "GOIABADA" in normalized_name:
            code = "G"
        elif "DOCE" in normalized_name:
            code = "D"
        elif "REQUEIJ" in normalized_name:
            code = "R"
        elif "QUEIJO" in normalized_name:
            code = "Q"

        if not code:
            continue

        current = product_by_code.get(code)
        if current is None or (not current["active"] and row["active"]):
            product_by_code[code] = {
                "id": row["id"],
                "name": row["name"],
                "unit_price": float(row["price"] or 0),
                "active": bool(row["active"]),
            }

    missing_codes = [code for code in ("T", "G", "D", "Q", "R") if code not in product_by_code]
    if missing_codes:
        raise RuntimeError(f"produtos oficiais ausentes para os codigos: {', '.join(missing_codes)}")

    historical_product = connection.execute(
        "select id, name, price, active from Product where name = ? order by id asc limit 1",
        (HISTORICAL_CUSTOM_BOX_NAME,),
    ).fetchone()
    if historical_product is None:
        cursor = connection.execute(
            """
            insert into Product (name, category, unit, price, active, createdAt)
            values (?, ?, 'unidade', ?, 0, ?)
            """,
            (
                HISTORICAL_CUSTOM_BOX_NAME,
                HISTORICAL_PRODUCT_CATEGORY,
                round(BOX_PRICE_CUSTOM / ORDER_BOX_UNITS, 6),
                datetime.now(tz=timezone.utc).isoformat().replace("+00:00", "Z"),
            ),
        )
        historical_product = connection.execute(
            "select id, name, price, active from Product where id = ?",
            (int(cursor.lastrowid),),
        ).fetchone()
    elif abs(float(historical_product["price"] or 0) - BOX_PRICE_CUSTOM) > 0.00001:
        connection.execute(
            "update Product set price = ?, category = ?, active = 0 where id = ?",
            (BOX_PRICE_CUSTOM, HISTORICAL_PRODUCT_CATEGORY, int(historical_product["id"])),
        )
        historical_product = connection.execute(
            "select id, name, price, active from Product where id = ?",
            (int(historical_product["id"]),),
        ).fetchone()

    product_by_code[HISTORICAL_CUSTOM_BOX_CODE] = {
        "id": historical_product["id"],
        "name": historical_product["name"],
        "unit_price": float(historical_product["price"] or 0),
        "active": bool(historical_product["active"]),
    }

    return product_by_code


def normalize_historical_custom_box_items(connection: sqlite3.Connection, product_id: int) -> int:
    rows = connection.execute(
        """
        select oi.id, oi.quantity, oi.total
        from OrderItem oi
        join "Order" o on o.id = oi.orderId
        where oi.productId = ?
          and o.notes like ?
        """,
        (product_id, f"{IMPORT_TAG} key=%"),
    ).fetchall()

    updated = 0
    for row in rows:
        quantity = int(row["quantity"] or 0)
        if quantity <= 0:
            continue
        next_quantity = quantity
        if quantity % ORDER_BOX_UNITS == 0:
            next_quantity = max(quantity // ORDER_BOX_UNITS, 1)
        next_total = round(float(row["total"] or 0), 2)
        next_unit_price = round(next_total / next_quantity, 6) if next_quantity > 0 else BOX_PRICE_CUSTOM
        if next_quantity == quantity and abs(next_unit_price - float(BOX_PRICE_CUSTOM)) < 0.00001:
            continue
        connection.execute(
            "update OrderItem set quantity = ?, unitPrice = ?, total = ? where id = ?",
            (next_quantity, next_unit_price, next_total, int(row["id"])),
        )
        updated += 1

    return updated


def load_customer_lookup(connection: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    rows = connection.execute("select id, name, deletedAt from Customer order by id asc").fetchall()
    for row in rows:
        key = normalize_lookup(row["name"])
        if not key:
            continue
        current = lookup.get(key)
        candidate = {
            "id": row["id"],
            "name": row["name"],
            "deleted_at": row["deletedAt"],
        }
        if current is None:
            lookup[key] = candidate
            continue
        if current["deleted_at"] is not None and row["deletedAt"] is None:
            lookup[key] = candidate
    return lookup


def load_existing_import_keys(connection: sqlite3.Connection) -> set[str]:
    rows = connection.execute('select notes from "Order" where notes like ?', (f"{IMPORT_TAG} key=%",)).fetchall()
    keys: set[str] = set()
    for row in rows:
        note = row["notes"] or ""
        prefix = f"{IMPORT_TAG} key="
        if not note.startswith(prefix):
            continue
        payload = note[len(prefix) :]
        key = payload.split(" origem=", 1)[0].strip()
        if key:
            keys.add(key)
    return keys


def load_existing_import_resolutions(connection: sqlite3.Connection) -> dict[str, ResolutionEntry]:
    rows = connection.execute('select notes from "Order" where notes like ?', (f"{IMPORT_TAG} key=%",)).fetchall()
    resolutions: dict[str, ResolutionEntry] = {}

    for row in rows:
        note = row["notes"] or ""
        prefix = f"{IMPORT_TAG} key="
        if not note.startswith(prefix):
            continue

        payload = note[len(prefix) :]
        import_key, separator, remainder = payload.partition(" origem=")
        if not separator or not import_key:
            continue

        _, boxes_separator, boxes_payload = remainder.partition(" caixas=")
        if not boxes_separator or not boxes_payload:
            continue

        for marker in (" ignorados=", " resolucao="):
            if marker in boxes_payload:
                boxes_payload = boxes_payload.split(marker, 1)[0]

        resolution_rows = [
            normalize_text(part)
            for part in re.split(r"\s*\|\s*", boxes_payload)
            if normalize_text(part)
        ]
        if not resolution_rows:
            continue

        resolutions[import_key.strip()] = ResolutionEntry(
            import_key=import_key.strip(),
            resolution_rows=resolution_rows,
            resolution_notes="[AUTO] carregada das notas de pedido importado",
        )

    return resolutions


def cleanup_duplicate_imported_orders(connection: sqlite3.Connection) -> dict[str, int]:
    rows = connection.execute(
        """
        select notes, group_concat(id) as ids, count(*) as total
        from "Order"
        where notes like ?
        group by notes
        having count(*) > 1
        """,
        (f"{IMPORT_TAG} key=%",),
    ).fetchall()

    removed_orders = 0
    removed_items = 0
    removed_payments = 0
    removed_movements = 0

    for row in rows:
        ids = sorted(int(value) for value in str(row["ids"]).split(",") if value)
        duplicate_ids = ids[1:]
        if not duplicate_ids:
            continue

        placeholders = ",".join("?" for _ in duplicate_ids)
        removed_items += connection.execute(
            f"delete from OrderItem where orderId in ({placeholders})",
            duplicate_ids,
        ).rowcount
        removed_payments += connection.execute(
            f"delete from Payment where orderId in ({placeholders})",
            duplicate_ids,
        ).rowcount
        removed_movements += connection.execute(
            f"delete from InventoryMovement where orderId in ({placeholders})",
            duplicate_ids,
        ).rowcount
        removed_orders += connection.execute(
            f'delete from "Order" where id in ({placeholders})',
            duplicate_ids,
        ).rowcount

    return {
        "removed_orders": removed_orders,
        "removed_items": removed_items,
        "removed_payments": removed_payments,
        "removed_movements": removed_movements,
    }


def ensure_customer(
    connection: sqlite3.Connection,
    customer_lookup: dict[str, dict[str, Any]],
    customer_name: str,
    created_at: str,
) -> tuple[int, bool]:
    customer_key = normalize_lookup(customer_name)
    existing = customer_lookup.get(customer_key)
    if existing is not None:
        return int(existing["id"]), False

    cursor = connection.execute(
        """
        insert into Customer (
          name, firstName, lastName, email, phone, address, addressLine1, addressLine2,
          neighborhood, city, state, postalCode, country, placeId, lat, lng, deliveryNotes,
          createdAt, deletedAt
        ) values (?, null, null, null, null, null, null, null, null, null, null, null, 'Brasil', null, null, null, ?, ?, null)
        """,
        (customer_name, IMPORT_TAG, created_at),
    )
    customer_id = int(cursor.lastrowid)
    customer_lookup[customer_key] = {"id": customer_id, "name": customer_name, "deleted_at": None}
    return customer_id, True


def import_orders(
    connection: sqlite3.Connection,
    plan: list[dict[str, Any]],
    product_by_code: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    customer_lookup = load_customer_lookup(connection)
    existing_import_keys = load_existing_import_keys(connection)

    imported_orders = 0
    created_customers = 0
    skipped_existing = 0
    imported_boxes = 0
    imported_units = 0
    imported_dates: list[str] = []

    for entry in plan:
        if entry["import_key"] in existing_import_keys:
            skipped_existing += 1
            continue

        customer_id, customer_created = ensure_customer(
            connection,
            customer_lookup,
            entry["customer_name"],
            entry["scheduled_at"],
        )
        if customer_created:
            created_customers += 1

        cursor = connection.execute(
            """
            insert into "Order" (
              customerId, status, subtotal, discount, total, notes, scheduledAt, createdAt
            ) values (?, 'ENTREGUE', ?, 0, ?, ?, ?, ?)
            """,
            (
                customer_id,
                entry["subtotal"],
                entry["subtotal"],
                entry["notes"],
                entry["scheduled_at"],
                entry["scheduled_at"],
            ),
        )
        order_id = int(cursor.lastrowid)

        for code, quantity in sorted(entry["counts"].items()):
            product = product_by_code[code]
            unit_price = float(product["unit_price"])
            total = round(unit_price * int(quantity), 2)
            connection.execute(
                """
                insert into OrderItem (orderId, productId, quantity, unitPrice, total)
                values (?, ?, ?, ?, ?)
                """,
                (order_id, product["id"], int(quantity), unit_price, total),
            )
            imported_units += int(quantity)

        connection.execute(
            """
            insert into Payment (orderId, amount, method, status, paidAt, dueDate, providerRef)
            values (?, ?, 'pix', 'PAGO', ?, null, ?)
            """,
            (order_id, entry["subtotal"], entry["scheduled_at"], f"{IMPORT_TAG}:{entry['import_key']}"),
        )

        existing_import_keys.add(entry["import_key"])
        imported_orders += 1
        imported_boxes += int(entry["box_count"])
        imported_dates.append(entry["iso_date"])

    return {
        "imported_orders": imported_orders,
        "created_customers": created_customers,
        "skipped_existing": skipped_existing,
        "imported_boxes": imported_boxes,
        "imported_units": imported_units,
        "date_range": {
            "from": min(imported_dates) if imported_dates else None,
            "to": max(imported_dates) if imported_dates else None,
        },
    }


def build_summary(
    workbook_path: Path,
    importable: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    reasons: Counter[str],
    excluded_reasons: Counter[str],
    apply_summary: dict[str, Any],
    cleanup_summary: dict[str, int],
    normalization_summary: dict[str, int],
    dry_run: bool,
) -> dict[str, Any]:
    return {
        "workbook": str(workbook_path),
        "mode": "dry-run" if dry_run else "apply",
        "generated_at": datetime.now(tz=timezone.utc).isoformat().replace("+00:00", "Z"),
        "totals": {
            "grouped_orders": len(importable) + len(skipped) + len(excluded),
            "importable_orders": len(importable),
            "skipped_orders": len(skipped),
            "excluded_orders": len(excluded),
            "skipped_reasons": dict(sorted(reasons.items())),
            "excluded_reasons": dict(sorted(excluded_reasons.items())),
        },
        "cleanup": cleanup_summary,
        "normalization": normalization_summary,
        "apply": apply_summary,
        "sample_importable": [
            {
                "iso_date": entry["iso_date"],
                "customer_name": entry["customer_name"],
                "time_hint": entry["time_hint"],
                "box_count": entry["box_count"],
                "counts": entry["counts"],
                "subtotal": entry["subtotal"],
                "raw_rows": entry["raw_rows"],
            }
            for entry in importable[:20]
        ],
        "sample_excluded": excluded[:20],
        "sample_skipped": skipped[:20],
    }


def write_pending_csv(skipped: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "import_key",
                "sheet",
                "iso_date",
                "time_hint",
                "customer_name",
                "raw_rows",
                "invalid_reasons",
                "invalid_rows",
                "resolution_rows",
                "resolution_notes",
            ],
        )
        writer.writeheader()
        for entry in skipped:
            writer.writerow(
                {
                    "import_key": entry.get("import_key", ""),
                    "sheet": entry["sheet"],
                    "iso_date": entry["iso_date"],
                    "time_hint": entry["time_hint"] or "",
                    "customer_name": entry["customer_name"],
                    "raw_rows": " | ".join(entry["raw_rows"]),
                    "invalid_reasons": " | ".join(
                        sorted({row["reason"] for row in entry.get("invalid_rows", []) if row.get("reason")})
                    ),
                    "invalid_rows": json.dumps(entry.get("invalid_rows", []), ensure_ascii=False),
                    "resolution_rows": "",
                    "resolution_notes": "",
                }
            )


def write_excluded_csv(excluded: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "import_key",
                "sheet",
                "iso_date",
                "time_hint",
                "customer_name",
                "raw_rows",
                "excluded_reasons",
                "excluded_rows",
            ],
        )
        writer.writeheader()
        for entry in excluded:
            writer.writerow(
                {
                    "import_key": entry.get("import_key", ""),
                    "sheet": entry["sheet"],
                    "iso_date": entry["iso_date"],
                    "time_hint": entry["time_hint"] or "",
                    "customer_name": entry["customer_name"],
                    "raw_rows": " | ".join(entry["raw_rows"]),
                    "excluded_reasons": " | ".join(
                        sorted({row["reason"] for row in entry.get("ignored_rows", []) if row.get("reason")})
                    ),
                    "excluded_rows": json.dumps(entry.get("ignored_rows", []), ensure_ascii=False),
                }
            )


def write_resolution_csv(
    skipped: list[dict[str, Any]],
    existing_resolutions: dict[str, ResolutionEntry],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "import_key",
                "sheet",
                "iso_date",
                "time_hint",
                "customer_name",
                "raw_rows",
                "invalid_reasons",
                "invalid_rows",
                "resolution_rows",
                "resolution_notes",
            ],
        )
        writer.writeheader()
        for entry in skipped:
            current = existing_resolutions.get(entry.get("import_key", ""))
            writer.writerow(
                {
                    "import_key": entry.get("import_key", ""),
                    "sheet": entry["sheet"],
                    "iso_date": entry["iso_date"],
                    "time_hint": entry["time_hint"] or "",
                    "customer_name": entry["customer_name"],
                    "raw_rows": " | ".join(entry["raw_rows"]),
                    "invalid_reasons": " | ".join(
                        sorted({row["reason"] for row in entry.get("invalid_rows", []) if row.get("reason")})
                    ),
                    "invalid_rows": json.dumps(entry.get("invalid_rows", []), ensure_ascii=False),
                    "resolution_rows": " | ".join(current.resolution_rows) if current else "",
                    "resolution_notes": current.resolution_notes or "" if current else "",
                }
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Importa pedidos legados do XLS para o banco local atual.")
    parser.add_argument(
        "--workbook",
        default="/Users/gui/Desktop/@QUEROBROAPP DOCS/antigos/QUERO BROA (1).xlsx",
        help="Caminho do arquivo XLSX legado.",
    )
    parser.add_argument(
        "--database",
        default="apps/api/prisma/dev.db",
        help="Caminho do banco SQLite atual.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Aplica a importacao. Sem isso, roda apenas em dry-run.",
    )
    parser.add_argument(
        "--summary-out",
        default="output/spreadsheet/legacy-order-import-summary.json",
        help="Arquivo JSON de resumo.",
    )
    parser.add_argument(
        "--pending-csv-out",
        default="output/spreadsheet/legacy-order-import-pending.csv",
        help="Arquivo CSV com pedidos pendentes de revisao manual.",
    )
    parser.add_argument(
        "--excluded-csv-out",
        default="output/spreadsheet/legacy-order-import-excluded.csv",
        help="Arquivo CSV com pedidos excluidos do escopo por regra de negocio.",
    )
    parser.add_argument(
        "--resolution-csv",
        default="output/spreadsheet/legacy-order-import-resolutions.csv",
        help="CSV com resolucoes manuais por import_key. Coluna resolution_rows usa caixas separadas por |.",
    )
    parser.add_argument(
        "--cleanup-duplicates",
        action="store_true",
        help="Remove pedidos importados duplicados mantendo o menor id por nota de importacao.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook)
    database_path = Path(args.database)
    summary_path = Path(args.summary_out)
    pending_csv_path = Path(args.pending_csv_out)
    excluded_csv_path = Path(args.excluded_csv_out)
    resolution_csv_path = Path(args.resolution_csv)

    if not workbook_path.exists():
        raise SystemExit(f"workbook nao encontrado: {workbook_path}")
    if not database_path.exists():
        raise SystemExit(f"banco nao encontrado: {database_path}")

    rows = load_rows(workbook_path)
    groups = group_rows(rows)

    summary_path.parent.mkdir(parents=True, exist_ok=True)

    apply_summary: dict[str, Any] = {
        "imported_orders": 0,
        "created_customers": 0,
        "skipped_existing": 0,
        "imported_boxes": 0,
        "imported_units": 0,
        "date_range": {"from": None, "to": None},
    }
    cleanup_summary = {
        "removed_orders": 0,
        "removed_items": 0,
        "removed_payments": 0,
        "removed_movements": 0,
    }
    normalization_summary = {
        "normalized_historical_custom_box_items": 0,
    }

    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row

    try:
        file_resolutions = load_resolutions(resolution_csv_path)
        persisted_resolutions = load_existing_import_resolutions(connection)
        resolutions = {**persisted_resolutions, **file_resolutions}
        importable, skipped, excluded, reasons, excluded_reasons = build_import_plan_with_resolutions(
            groups, resolutions
        )
        product_by_code = fetch_products(connection)
        if args.cleanup_duplicates or args.apply:
            with connection:
                if args.cleanup_duplicates:
                    cleanup_summary = cleanup_duplicate_imported_orders(connection)
                normalization_summary["normalized_historical_custom_box_items"] = (
                    normalize_historical_custom_box_items(
                        connection, int(product_by_code[HISTORICAL_CUSTOM_BOX_CODE]["id"])
                    )
                )
                if args.apply:
                    apply_summary = import_orders(connection, importable, product_by_code)

        summary = build_summary(
            workbook_path=workbook_path,
            importable=importable,
            skipped=skipped,
            excluded=excluded,
            reasons=reasons,
            excluded_reasons=excluded_reasons,
            apply_summary=apply_summary,
            cleanup_summary=cleanup_summary,
            normalization_summary=normalization_summary,
            dry_run=not args.apply,
        )
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        write_pending_csv(skipped, pending_csv_path)
        write_excluded_csv(excluded, excluded_csv_path)
        write_resolution_csv(skipped, resolutions, resolution_csv_path)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    finally:
        connection.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
