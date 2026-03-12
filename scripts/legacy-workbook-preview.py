from __future__ import annotations

import json
import re
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MONTH_SHEETS = {"AGOSTO", "SETEMBRO", "OUTUBRO", "JANEIRO", "FEVEREIRO"}
TIME_PREFIX_RE = re.compile(r"^(?P<time>\d{1,2}h(?:\d{2})?)\s*-\s*(?P<body>.+)$", re.IGNORECASE)


@dataclass
class LegacyOrderPreview:
    sheet: str
    iso_date: str
    column: int
    row: int
    raw_value: str
    customer_name: str
    time_hint: str | None
    order_hint: str | None


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def split_legacy_entry(raw: str) -> tuple[str, str | None, str | None]:
    normalized = normalize_text(raw)
    if not normalized:
        return "", None, None

    time_hint = None
    body = normalized
    time_match = TIME_PREFIX_RE.match(normalized)
    if time_match:
        time_hint = normalize_text(time_match.group("time"))
        body = normalize_text(time_match.group("body"))

    if " - " in body:
        customer_name, order_hint = body.split(" - ", 1)
        return normalize_text(customer_name), time_hint, normalize_text(order_hint)

    return body, time_hint, None


def iter_month_previews(workbook_path: Path) -> list[LegacyOrderPreview]:
    wb = load_workbook(workbook_path, data_only=True)
    previews: list[LegacyOrderPreview] = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTH_SHEETS:
            continue

        ws = wb[sheet_name]
        for col in range(1, ws.max_column + 1):
            date_value = ws.cell(1, col).value
            if date_value is None:
                continue
            if hasattr(date_value, "date"):
                iso_date = date_value.date().isoformat()
            else:
                iso_date = normalize_text(str(date_value))

            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row, col).value
                if cell_value is None:
                    continue
                raw_value = normalize_text(str(cell_value))
                if not raw_value:
                    continue

                customer_name, time_hint, order_hint = split_legacy_entry(raw_value)
                previews.append(
                    LegacyOrderPreview(
                        sheet=sheet_name,
                        iso_date=iso_date,
                        column=col,
                        row=row,
                        raw_value=raw_value,
                        customer_name=customer_name,
                        time_hint=time_hint,
                        order_hint=order_hint,
                    )
                )

    return previews


def build_summary(previews: list[LegacyOrderPreview]) -> dict[str, Any]:
    by_sheet = Counter(preview.sheet for preview in previews)
    with_time = sum(1 for preview in previews if preview.time_hint)
    with_hint = sum(1 for preview in previews if preview.order_hint)
    top_hints = Counter(preview.order_hint for preview in previews if preview.order_hint).most_common(20)

    return {
      "total_entries": len(previews),
      "entries_by_sheet": dict(sorted(by_sheet.items())),
      "entries_with_time_hint": with_time,
      "entries_with_order_hint": with_hint,
      "top_order_hints": [{"hint": hint, "count": count} for hint, count in top_hints],
      "sample": [asdict(preview) for preview in previews[:40]],
    }


def main() -> int:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if not workbook_path or not workbook_path.exists():
        print("usage: python3 scripts/legacy-workbook-preview.py /path/to/workbook.xlsx", file=sys.stderr)
        return 1

    previews = iter_month_previews(workbook_path)
    summary = build_summary(previews)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
